import requests
import re
import sys
import urllib


import asyncio
from pyppeteer import launch
from pyquery import PyQuery as pq
from bs4 import BeautifulSoup
import openpyxl
import time

async def one(url):
    global m
    browser = await launch()
    page = await browser.newPage()
    await page.goto(url)
    html = pq(await page.content())
    #print(type(html))
    soup = BeautifulSoup(str(html), 'html5lib')
    data_list = []
    for idx, tr in enumerate(soup.find_all('tr')):
        if idx != 0:
            tds = tr.find_all('td')
            data_list.append({
                '日期': tds[0].contents[0],
                'AQI': tds[1].contents[0],
                'PM2.5': tds[3].contents[0],
                'PM10': tds[4].contents[0],
                'SO2': tds[5].contents[0],
                'CO': tds[6].contents[0],
                'NO2': tds[7].contents[0],
                '03': tds[8].contents[0]
            })
    #print(data_list)
    for i in data_list:
        ws.cell(m,1,str(i['日期']))
        ws.cell(m, 2, str(i['AQI']))
        ws.cell(m, 3, str(i['PM2.5']))
        ws.cell(m, 4, str(i['PM10']))
        ws.cell(m, 5, str(i['SO2']))
        ws.cell(m, 6, str(i['CO']))
        ws.cell(m, 7, str(i['NO2']))
        ws.cell(m, 8, str(i['03']))
        m=m+1

    await browser.close()





def datte():
    cc=[]
    for e in range(2015,2019):
        for j in range(1,13):
            p=j
            if len(str(p))==1:
                p='0'+str(p)
                u = str(e)+p
            else:
                u=str(e)+str(p)
            cc.append(u)
            u=''
    return cc







if __name__=="__main__":
    wb = openpyxl.Workbook()
    ws = wb.active
    m=1
    ddd=datte()
    for que in ddd:
        url = ""
        print(url)
        asyncio.get_event_loop().run_until_complete(one(url))
        time.sleep(10)
    wb.save('example.xlsx')

